from openpyxl import load_workbook

def extract_text_from_xlsx(file_path, max_rows=20, max_sheets=3):
    """
    Extract text from Excel file with limits
    """
    text = []

    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        
        # Limit number of sheets processed
        sheets_to_process = list(wb.worksheets)[:max_sheets]

        for sheet_idx, sheet in enumerate(sheets_to_process):
            text.append(f"\n=== Sheet: {sheet.title} ===")

            rows = list(sheet.iter_rows(values_only=True, max_row=max_rows + 1))
            if not rows:
                text.append("[Empty sheet]")
                continue

            # Headers
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            if any(headers):
                text.append("Columns: " + ", ".join(h for h in headers if h))

            # Data rows
            data_rows = 0
            for row in rows[1:]:
                if not any(row):
                    continue

                row_text = [str(cell).strip() if cell is not None else "" for cell in row]
                if any(row_text):
                    text.append(" | ".join(cell for cell in row_text if cell))
                    data_rows += 1
            
            if data_rows == 0:
                text.append("[No data rows]")
        
        # Note if there are more sheets
        if len(wb.worksheets) > max_sheets:
            text.append(f"\n[Note: File has {len(wb.worksheets)} sheets, only first {max_sheets} shown]")

        wb.close()

    except Exception as e:
        return f"[ERROR reading Excel file: {str(e)}]"

    return "\n".join(text)